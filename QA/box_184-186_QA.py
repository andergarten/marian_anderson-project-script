
# coding: utf-8

# In[21]:


import os
import shutil
import subprocess
from openpyxl import *
from distutils.dir_util import copy_tree

def get_ws(wb_path):
    '''get workbook sheet
    '''
    # open xlsx with ark id; use data_only mode to exclude the formula
    wb = load_workbook(wb_path, data_only=True)
    # get the first worksheet
    ws = wb["Sheet1"]
    return ws

def get_file_list(path):
    '''get a list of sorted file name
    '''
    # get original file_list
    file_list = os.listdir(path)
    # sort file_list in alphabetic order
    file_list.sort()
    #print (file_list)
    return file_list

# def initialization(path, drive_letter):
#     '''This method is to change current work folder, get data worksheet, 
#     copy folder content, and get file list in folder
#     '''
#     os.chdir(path)
#     # get worksheet
#     ws = get_ws(drive_letter)
#     # copy current folder
#     copy_folder(path)
#     # get file list
#     file_list = get_file_list(path)
    
#     return file_list

def get_folder_row(ws, folder_name):
    '''
    '''
    # initialize variable
    folder_row = []
    # iterate folder name column in xlsx
    for row in ws.iter_rows(min_row = 368, min_col = 7, max_col = 7, max_row = 1070):
        for cell in row:
            # find rows where current folder is, and record row number
            #print(str(cell.value), folder_name)
            if str(cell.value) == folder_name:
                folder_row.append(cell.row)
    #print(folder_row)
    return folder_row

def get_reference(ws, folder_name):
    '''get a list of image number of reference shot
    '''
    # initialize answer
    answer = 'n'
    while answer == 'n':
        # get folder_row
        folder_row = get_folder_row(ws, folder_name)
        item_num = ws.cell(row = folder_row[len(folder_row)-1], column = 8).value
        
        # initialize reference list
        reference = []
        # get first image number of reference shot
        num = input("Please give me the image number for the first reference shot.")
        # add it to list
        reference.append(int(num))
        # get all image numbers
        while(num != "n"):
            print(reference)
            num = input("Please give me the image number for the next reference shot. " + \
                "If there is no reference shot, type n.")
            if num == "n":
                break
            else:
                reference.append(int(num))
                
        # print current reference list        
        print('current reference image list: ')
        print(reference)

        # see if reference shot number match item number
        if len(reference) != item_num:
            print('There are ' + str(item_num) + ' items in this folder, but now '\
                  + str(len(reference)) + ' reference shot numbers were entered.')
            answer = input('Are you sure to continue? y for continue, n for re-enter.')
        else:
            answer = 'y'
    
    return reference

def get_ark(ws, folder_name, item_number):
    '''This method is to get folder rows
    '''
    # get folder_row
    folder_row = get_folder_row(ws, folder_name)
    # iterate rows where current folder is
    for row in folder_row:
        # find where current item is
        if ws.cell(row = row, column = 8).value == item_number:
            # extract the ark id
            ark_id = ws.cell(row = row, column = 4).value
    
    return ark_id

def copy_folder(path):
    if not os.path.exists(path + ' - Copy'):
        copy_tree(path, path + ' - Copy')

def rename_183_plus(path, folder_name, wb_path):
    '''This method is to rename file in boxes after box 183
    for method refold_and_renme to use
    '''
    os.chdir(path)
    # get worksheet
    ws = get_ws(wb_path)
    # copy current folder
    copy_folder(path)
    # get file list
    file_list = get_file_list(path)

    # initialize variables
    item_number = 0
    # page number
    image_number = 0
    # total image number
    total_number = 1
    ark_id = ""
    # get reference shot list
    reference = get_reference(ws, folder_name)
    
    # iterate over file list in folder
    for file_name in file_list:
        # pass CaptureOne folder and .DS_Store file and folder already with ark_id
        if file_name != "CaptureOne" and file_name != ".DS_Store" and ".tif" in file_name:
            # if it is a sample shot
            if total_number in reference:
                # start rename a new item
                item_number += 1
                image_number = 0
                total_number += 1
                # get ark id
                ark_id = get_ark(ws, folder_name, item_number)      
                # renme the sample file
                os.rename(file_name, ark_id + "_body0000testref1.tif")

            else:
                # if it is a normal image, rename it and increase image number and total number
                image_number += 1
                total_number += 1
                os.rename(file_name, ark_id + "_body000" + str(image_number) + ".tif")
                
def rename_180_plus(path, folder_name, wb_path, sample_shot):
    '''This method is to rename file in boxes after box 180
    for method refold_and_renme to use
    '''
    os.chdir(path)
    # get worksheet
    ws = get_ws(wb_path)
    # copy current folder
    copy_folder(path)
    # get file list
    file_list = get_file_list(path)

    # initialize variables
    item_number = 0
    image_number = 0
    ark_id = ""

    # iterate over file list in folder
    for file_name in file_list:
        #print (os.path.getsize(path + "\\" + file_name))
        # pass CaptureOne folder and .DS_Store file and folder already with ark id
        if file_name != "CaptureOne" and file_name != ".DS_Store" and ".tif" in file_name:
            # if it is a sample picture
            if os.path.getsize(file_name) == os.path.getsize(sample_shot):
                # start rename a new item
                item_number += 1
                image_number = 0
                # get ark id
                ark_id = get_ark(ws, folder_name, item_number)      
                # renme the sample file
                os.rename(file_name, ark_id + "_body0000testref1.tif")

            else:
                #print (file_name)
                # if it is a normal image, rename it and increase image number
                image_number += 1
                os.rename(file_name, ark_id + "_body000" + str(image_number) + ".tif")

def refold_and_rename(path):
    '''This method is to refold and rename file in boxes
    '''
    # iterate over file list in folder
    os.chdir(path)
    for file_name in os.listdir('.'):
        # pass CaptureOne folder and .DS_Store file
        # only proceed if this folder isn't refolded
        if file_name != "CaptureOne" and file_name != ".DS_Store" and ".tif" in file_name:
            # obtain ark id
            ark_id = ''
            for charactor in file_name:
                if charactor != '_':
                    ark_id += charactor
                else:
                    break

            # make new folder with ark id
            new_path = ark_id
            if not os.path.exists(new_path):
                os.makedirs(new_path)

            # move files and rename
            old_file = file_name
            
            # if it is not a sample image, rename
            if "testref1" not in file_name:
                # if it is a regular page, keep last 8 digits
                if 'a.' not in file_name:
                    new_file = new_path + '\\' + file_name[-8:]
                    #new_file = new_path + '/' + file_name[-8:]
                # if it is named irregularly, keep last 9 digits
                else:
                    new_file = new_path + '\\' + file_name[-9:]
                    #new_file = new_path + '/' + file_name[-9:]       
            # otherwise, do not rename
            else:
                new_file = new_path + '\\' + file_name
                #new_file = new_path + '/' + file_name
            
            # move to new folder named as ark_id    
            shutil.move(old_file, new_file)

# def get_page_number(path, ark_id, page_number):
#     '''This method is to get actual page number and compare with that
#     in the record to see if it is matched
#     '''

def create_spreadsheet(path, ark_id, page_number, tem_path):
    '''The method is to create a spreadsheet everytime
    after we do the QA for each item
    '''
    
    # copy spreadsheet template file to folder path
    shutil.copyfile(tem_path, path + "\\Copy of Marian Anderson spreadsheet template.xlsx")

    # change cwd to folder_path
    os.chdir(path)
        
    # open template
    template = load_workbook('Copy of Marian Anderson spreadsheet template.xlsx')

    # get the first worksheet
    sheet1 = template["Structural Metadata"]

    '''
    # get page number for each ark_id item
    page_number = get_page_number(box_number, folder_number, ark_id, page_number)
    '''

    # use loop to fill sheet
    for page in range(1, page_number+1):

        # fill "visible page" field
        sheet1['B'+ str(3 + page)] = page;

        # fill "filename" field
        page_str = str(page)
        if len(page_str) == 1:
            sheet1['D'+ str(3 + page)] = '000' + str(page) + '.tif'
        if len(page_str) == 2:
            sheet1['D'+ str(3 + page)] = '00' + str(page) + '.tif'
        if len(page_str) == 3:
            sheet1['D'+ str(3 + page)] = '0' + str(page) + '.tif'
        if len(page_str) == 4:
            sheet1['D'+ str(3 + page)] = str(page) + '.tif'

    # save updated file
    template.save('81431' + ark_id + '.xlsx')

    # remove extra template file
    if os.path.exists('Copy of Marian Anderson spreadsheet template.xlsx'):
        os.remove('Copy of Marian Anderson spreadsheet template.xlsx')

def refine_arklist(path):
    '''This method is to refine the file to eliminate the items that
    have already been checked
    '''
    
    # get existing xlsx file in a list
    xlsx_list = [x for x in os.listdir(path) if os.path.isfile(os.path.join(path,x))]
    # loop over the ark_id folders
    ark_list = [a for a in os.listdir(path) if not os.path.isfile(os.path.join(path,a))]

    #update ark_id_list
    for xlsx_name in xlsx_list:
        for ark_id in ark_list:
            #if spreadsheet has already been created, skip it
            if ark_id in xlsx_name:
                #print (xlsx_name, ark_id)
                ark_list.remove(ark_id)

    return ark_list

def qa(path, tem_path):
    '''This method is to ask students to do QA and give feedback.
    QA instruction
    '''

    # use refined arklist to skip item which has already been checked
    ark_list = refine_arklist(path)
   
    for ark_id in ark_list:

        # exclude non-ark_id folder
        if "CaptureOne" not in ark_id:

            # ask student to do QA
            qa_result = input("Does " + ark_id + " pass QA? y for pass, n for fail.\n")

            while (qa_result != "y" and qa_result != "n"):
                print (qa_result)
                # ask for input again
                qa_result = input("Wrong input. y for pass, n for fail.\n")
                
            # if pass, create spreadsheet
            if qa_result == "y":
                # tell student what to do
                print("Please put ‘Y’ into ‘QA Pass’ field.")
                # ask for page number
                page_number = int(input("What is the page number?\n"))
                # create a spreadsheet
                create_spreadsheet(path, ark_id, page_number, tem_path)

            # if fail, 
            elif qa_result == "n":
                # tell student what to do
                print("Please put ‘N’ into ‘QA Fail’ field and fill in reasons. Email Craig with description of problem.")
                # ask for page number
                page_number = int(input("What is the correct page number?\n"))
                # create a spreadsheet
                create_spreadsheet(path, ark_id, page_number, tem_path)

def main():
    
    # install openpyxl package
    subprocess.check_call(["python", '-m', 'pip', 'install', 'openpyxl'])

    # # for test
    # drive_letter = 'T'
    # box_number = '184'
    # folder_number = '8632'
    # path = ''
    
    # ask for drive number
    drive_letter = input('What is the drive letter(in capital) for sceti-completed on your computer?\n')
    while True:
        # ask for box number
        box_number = str(input('Give me a box number.\n'))
        # get folder number
        folder_number = str(input('Give me a 4-digit folder number.\n'))
        # target folder
        path = drive_letter + r":\MarianAnderson\mscoll200_box" + box_number + "\\folder0" + folder_number
        #path = r"T:\MarianAnderson\mscoll200_box" + box_number + "\\folder0" + folder_number

        # address validation
        while not os.path.exists(path):
            print("No such folder. Please re-enter.")
            #drive_letter = input('What is the drive letter(in capital) for sceti-completed on your computer?\n')
            box_number = str(input('Give me a box number.\n'))
            foler_number = str(input('Give me a 4-digit folder number.\n'))
            path = drive_letter + r":\MarianAnderson\mscoll200_box" + box_number + "\\folder0" + folder_number
            #path = r"T:\MarianAnderson\mscoll200_box" + box_number + "\\folder0" + folder_number

    # .   # another way of input validation
    #     # initialize 
    #     isExist = False
    #     while not isExist:
    #         # try to open with path
    #         try:
    #             isExist = True
    #             open(path)

    #         except OSError as e:
    #             # label it not existing
    #             isExist = False
    #             # tell user
    #             print("No such folder. Please re-enter.")
    #             # re-enter information
    #             box_number = str(input('Give me a box number.\n'))
    #             folder_number = str(input('Give me a 4-digit folder number.\n'))
    #             path = drive_letter + r":\MarianAnderson\mscoll200_box" + box_number \
    #                    + "\\folder0" + folder_number

        wb_path = drive_letter + r":\MarianAnderson\Student Copy of Marian Anderson" + \
                ' Batch 1_ ARK IDs (Ms. Coll 200V, Boxes 180_189) .xlsx'
        #wb_path = r'/Users/zhiyuzhou/OneDrive/Penn/18_Summer/Work_on-campus/Script/QA/Student Copy of Marian Anderson Batch 1_ ARK IDs (Ms. Coll 200V, Boxes 180_189) .xlsx'
        tem_path = drive_letter + r":\MarianAnderson\Copy of Marian Anderson spreadsheet template.xlsx"
        #tem_path = r':/Users/zhiyuzhou/OneDrive/Penn/18_Summer/Work_on-campus/Script/QA/Copy of Marian Anderson spreadsheet template.xlsx'
        sample_shot = drive_letter + r":\MarianAnderson\folder08578_body0001.tif"

        # an extra rename step after box 183
        rename_183_plus(path, folder_number, wb_path)
        #rename_180_plus(path, folder_number, wb_path, sample_shot)

        # refold and rename
        refold_and_rename(path)
        print("Refolding and renaming are done...\n")

        # do QA
        qa(path, tem_path)
        print("QA and creating spreadsheets are done for folder" + \
              folder_number + "...\n")

        # if no error, delete the copy file
        #if os.path.exists(path + ' - Copy'):
            #shutil.rmtree(path + ' - Copy')

        # tell student to upload spreadsheet
        print("Now please upload spreadsheets in folder 0" + folder_number + " to Google Drive...\n")
                    
if __name__ == '__main__':
    main()


